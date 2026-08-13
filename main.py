import base64
import json
import os
import time
import io
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright
import requests

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

# ==================== KÖRNYEZETI VÁLTOZÓK (GITHUB SECRETS) ====================
TEYA_EMAIL = os.getenv("TEYA_EMAIL")
TEYA_PASSWORD = os.getenv("TEYA_PASSWORD")
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID", "0AP3HkLh_ANsVUk9PVA")
GDRIVE_SA_JSON_STR = os.getenv("GDRIVE_SA_JSON")

LOGIN_URL = "https://business.teya.com/?locale=hu"
COMPANY_ID = "24abe7e0-88fd-468a-afb6-fe362b0b79ca"
STORE_ID = "81e77008-fc09-4a82-9c78-b5939b24a202"

# ELMÚLT 14 NAP (2 HÉT) GYŰJTÉSE A NAPI ÉS NYERS FÁJLOKHOZ
DAYS_BACK = 14

OUTPUT_DIR = "./output"
GRAPHQL_URL = "https://customer-bff.teya.com/graphql"
# ==============================================================================


def get_drive_service():
    """Visszaadja a hitelesített Google Drive API szolgáltatást."""
    if not GDRIVE_SA_JSON_STR:
        return None
    try:
        sa_info = json.loads(GDRIVE_SA_JSON_STR)
        SCOPES = ["https://www.googleapis.com/auth/drive"]
        creds = service_account.Credentials.from_service_account_info(sa_info, scopes=SCOPES)
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        print(f"      [Drive Auth HIBA] {e}")
        return None


def download_master_file_from_drive(file_name, folder_id, dest_path):
    """Letölti a meglévő Master fájlt a Google Drive-ról, hogy ne törlődjön a korábbi adat."""
    service = get_drive_service()
    if not service or not folder_id:
        return False

    print(f"      [Drive] Keresés a korábbi {file_name} fájlhoz a felülíráshoz...")
    try:
        query = f"name='{file_name}' and '{folder_id}' in parents and trashed=false"
        results = service.files().list(
            q=query, spaces='drive', fields="files(id, name)", 
            supportsAllDrives=True, includeItemsFromAllDrives=True
        ).execute()
        
        items = results.get('files', [])
        if not items:
            print(f"      [Drive INFO] Nem létezik még a {file_name} fájl, a kód létrehozza újonnan.")
            return False

        file_id = items[0]['id']
        request = service.files().get_media(fileId=file_id)
        fh = io.FileIO(dest_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            
        print(f"      [Drive SIKER] A korábbi {file_name} sikeresen letöltve a szerkesztéshez!")
        return True
    except Exception as e:
        print(f"      [Drive LETÖLTÉS HIBA] {e}")
        return False


def upload_to_google_drive(file_path, file_name, folder_id):
    """Fájl feltöltése vagy FELÜLÍRÁSA Google Drive-on."""
    service = get_drive_service()
    if not service or not folder_id:
        return False

    print(f"      [Drive] Szinkronizálás: {file_name}...")
    try:
        query = f"name='{file_name}' and '{folder_id}' in parents and trashed=false"
        results = service.files().list(
            q=query, spaces='drive', fields="files(id, name)", 
            supportsAllDrives=True, includeItemsFromAllDrives=True
        ).execute()
        
        items = results.get('files', [])
        media = MediaFileUpload(
            file_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )

        if items:
            # Fájl felülírása (Update)
            file_id = items[0]['id']
            service.files().update(
                fileId=file_id, 
                media_body=media, 
                supportsAllDrives=True
            ).execute()
            print(f"      [Drive SIKER] Fájl felülírva (Frissítve).")
        else:
            # Új fájl létrehozása (Create)
            file_metadata = {"name": file_name, "parents": [folder_id]}
            service.files().create(
                body=file_metadata,
                media_body=media,
                fields="id",
                supportsAllDrives=True
            ).execute()
            print(f"      [Drive SIKER] Új fájl feltöltve.")
        return True

    except Exception as e:
        print(f"      [Drive HIBA] Nem sikerült a művelet ({file_name}): {e}")
        return False


def format_eur(amount):
    if amount is None: amount = 0.0
    formatted = f"{amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    return f"{formatted} €"


def create_summary_excel(file_path, summary, payout_date):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Összegzés"
        headers = ["Dátum", "Kifizetett Végösszeg", "Kezdeti egyenleg", "Tranzakciók bruttó", "Visszatérítések", "Chargeback", "Bankköltség / Díjak"]
        values = [
            payout_date, summary.get("kifizetes", "0,00 €"), summary.get("kezdeti_egyenleg", "0,00 €"),
            summary.get("tranzakciok", "0,00 €"), summary.get("visszateritesek", "0,00 €"),
            summary.get("chargeback", "0,00 €"), summary.get("dijak", "0,00 €")
        ]
        ws.append(headers)
        ws.append(values)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            if col_num == 2:
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                val_cell = ws.cell(row=2, column=col_num)
                val_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                val_cell.font = Font(bold=True)
        wb.save(file_path)
        return True
    except Exception as e:
        return False


def rebuild_master_summary_from_all_files(master_path, folder_id):
    """
    Összegyűjti az ÖSSZES Google Drive-on és helyi mappában található napi 'amberlyn_*_Osszegzes.xlsx'
    fájlt, és az összes visszamenőleges adatot belefoglalja a Master Összegzés fájlba!
    A 'Dátum' oszlopba a fájlnévből származó dátum kerül.
    """
    print("\n[MASTER ÖSSZEGZŐ ÚJRAÉPÍTÉSE AZ ÖSSZES KORÁBBI FÁJLBÓL]")
    service = get_drive_service()
    headers = ["Dátum", "Kifizetett Végösszeg", "Kezdeti egyenleg", "Tranzakciók bruttó", "Visszatérítések", "Chargeback", "Bankköltség / Díjak"]
    
    # dátum -> adatsor szótár
    all_summary_rows = {}

    # 1. Beolvasás a helyi mappából (a frissen lekért 2 hét adatai)
    if os.path.exists(OUTPUT_DIR):
        for f_name in os.listdir(OUTPUT_DIR):
            if f_name.startswith("amberlyn_") and f_name.endswith("_Osszegzes.xlsx") and f_name != "amberlyn_MASTER_Osszegzes.xlsx":
                date_part = f_name.replace("amberlyn_", "").replace("_Osszegzes.xlsx", "")
                f_path = os.path.join(OUTPUT_DIR, f_name)
                try:
                    wb = openpyxl.load_workbook(f_path)
                    ws = wb.active
                    if ws.max_row >= 2:
                        row_vals = [ws.cell(row=2, column=col).value for col in range(1, 8)]
                        row_vals[0] = date_part  # A dátum a fájlnévvel egyezik meg
                        all_summary_rows[date_part] = row_vals
                except Exception as e:
                    print(f"      [Helyi beolvasás HIBA] {f_name}: {e}")

    # 2. Beolvasás a Google Drive-ról (az ÖSSZES korábbi visszamenőleges napi összegző)
    if service and folder_id:
        try:
            query = f"name contains 'amberlyn_' and name contains '_Osszegzes.xlsx' and name != 'amberlyn_MASTER_Osszegzes.xlsx' and '{folder_id}' in parents and trashed=false"
            results = service.files().list(
                q=query, spaces='drive', fields="files(id, name)",
                supportsAllDrives=True, includeItemsFromAllDrives=True,
                pageSize=1000
            ).execute()

            drive_items = results.get('files', [])
            print(f"      [Drive Master Összegző] {len(drive_items)} db korábbi összegző fájl beazonosítva a Drive-on.")

            for item in drive_items:
                f_name = item['name']
                date_part = f_name.replace("amberlyn_", "").replace("_Osszegzes.xlsx", "")

                # Ha a helyi gyűjtésben még nincs meg (korábbi napok)
                if date_part not in all_summary_rows:
                    temp_path = os.path.join(OUTPUT_DIR, f"temp_{f_name}")
                    try:
                        request = service.files().get_media(fileId=item['id'])
                        fh = io.FileIO(temp_path, 'wb')
                        downloader = MediaIoBaseDownload(fh, request)
                        done = False
                        while not done:
                            status, done = downloader.next_chunk()

                        wb = openpyxl.load_workbook(temp_path)
                        ws = wb.active
                        if ws.max_row >= 2:
                            row_vals = [ws.cell(row=2, column=col).value for col in range(1, 8)]
                            row_vals[0] = date_part  # A dátum a fájlnévvel megegyező
                            all_summary_rows[date_part] = row_vals

                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                    except Exception as e:
                        print(f"      [Drive fájl letöltés HIBA] {f_name}: {e}")
                        if os.path.exists(temp_path):
                            os.remove(temp_path)

        except Exception as e:
            print(f"      [Drive listázás HIBA]: {e}")

    # 3. Master Összegző Excel fájl összeállítása dátum szerint rendezve
    wb_master = openpyxl.Workbook()
    ws_master = wb_master.active
    ws_master.title = "Master Összegzés"
    ws_master.append(headers)

    for col_num in range(1, len(headers) + 1):
        ws_master.cell(row=1, column=col_num).font = Font(bold=True)

    sorted_dates = sorted(all_summary_rows.keys())
    for d in sorted_dates:
        ws_master.append(all_summary_rows[d])

    wb_master.save(master_path)
    print(f"      [Master Összegző SIKER] Összesen {len(all_summary_rows)} nap adata összefűzve a Master fájlba!")


def append_to_master_raw(master_path, daily_raw_path, payout_date):
    if not os.path.exists(daily_raw_path):
        return
    try:
        daily_wb = openpyxl.load_workbook(daily_raw_path)
        daily_ws = daily_wb.active

        if not os.path.exists(master_path):
            master_wb = openpyxl.Workbook()
            master_ws = master_wb.active
            master_ws.title = "Master Tranzakciók"
            
            # Fejléc másolása + Egy Rendszer Dátum oszlop hozzáadása
            header_row = list(next(daily_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            header_row.append("API_Lekeres_Datuma")
            master_ws.append(header_row)
        else:
            master_wb = openpyxl.load_workbook(master_path)
            master_ws = master_wb.active

            # FELÜLÍRÁS: Visszafelé töröljük a régi sorokat, amik ehhez a naphoz tartoznak
            for row_idx in range(master_ws.max_row, 1, -1):
                if master_ws.cell(row=row_idx, column=master_ws.max_column).value == payout_date:
                    master_ws.delete_rows(row_idx, 1)

        # Új adatsorok hozzáfűzése a napi fájlból
        for row_idx, row_values in enumerate(daily_ws.iter_rows(values_only=True), start=1):
            if row_idx < 2:  # Fejlécet átugorjuk
                continue
            row_list = list(row_values)
            row_list.append(payout_date)  # Hozzáadjuk a dátumot az utolsó oszlophoz
            master_ws.append(row_list)

        master_wb.save(master_path)
    except Exception as e:
        print(f"      [Master Nyers HIBA]: {e}")


def generate_and_download_raw_excel(payout_id, save_path, headers):
    gen_payload = {
        "query": "query GenerateTransactionsReport($input: GenerateTransactionsReportInput!) {\n  generateTransactionsReport(input: $input) {\n    id\n  }\n}",
        "variables": {"input": {"storeId": STORE_ID, "payoutId": payout_id, "locale": "hu"}},
        "operationName": "GenerateTransactionsReport"
    }

    try:
        gen_res = requests.post(GRAPHQL_URL, headers=headers, json=gen_payload)
        if gen_res.status_code != 200: return False
        job_id = gen_res.json().get("data", {}).get("generateTransactionsReport", {}).get("id")
        if not job_id: return False

        poll_payload = {
            "query": "query GetSettlementReportsByJobId($jobId: ID!) {\n  getSettlementReportsByJobId(jobId: $jobId) {\n    jobId\n    reports {\n      documentCompleted\n      documentId\n      documentMimeType\n      currency\n      id\n      status\n    }\n    status\n  }\n}",
            "variables": {"jobId": job_id},
            "operationName": "GetSettlementReportsByJobId"
        }

        document_id = None
        for attempt in range(1, 12):
            time.sleep(1)
            poll_res = requests.post(GRAPHQL_URL, headers=headers, json=poll_payload)
            if poll_res.status_code == 200:
                job_data = poll_res.json().get("data", {}).get("getSettlementReportsByJobId", {})
                reports = job_data.get("reports", [])
                if reports and reports[0].get("documentCompleted"):
                    document_id = reports[0].get("documentId")
                    break

        if not document_id: return False

        dl_payload = {
            "query": "query DownloadDocument($input: DownloadDocumentInput!) {\n  downloadDocument(input: $input) {\n    documentBase64\n  }\n}",
            "variables": {"input": {"documentId": document_id, "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}},
            "operationName": "DownloadDocument"
        }

        dl_res = requests.post(GRAPHQL_URL, headers=headers, json=dl_payload)
        if dl_res.status_code == 200:
            b64_str = dl_res.json().get("data", {}).get("downloadDocument", {}).get("documentBase64")
            if b64_str:
                excel_bytes = base64.b64decode(b64_str)
                with open(save_path, "wb") as f:
                    f.write(excel_bytes)
                return True
    except Exception:
        pass
        
    return False


def fetch_all_settlements_in_chunks(start_date, end_date, headers):
    settlements_url = f"https://business.teya.com/reporting/v1/company/{COMPANY_ID}/settlements"
    all_items = []
    current_start = start_date

    while current_start < end_date:
        current_end = min(current_start + timedelta(days=25), end_date)
        start_str = current_start.strftime("%Y-%m-%d")
        end_str = current_end.strftime("%Y-%m-%d")
        
        offset = 0
        limit = 50
        while True:
            payload = {
                "begin_date": start_str,
                "end_date": end_str,
                "store_ids": [],
                "states": ["POSTPONED", "PARTIALLY_PAID", "PAID", "SCHEDULED"],
                "limit": limit,
                "offset": offset
            }
            res = requests.put(settlements_url, headers=headers, json=payload)
            if res.status_code != 200: break
            data = res.json().get("data", [])
            all_items.extend(data)
            if len(data) < limit: break
            offset += limit

        current_start = current_end + timedelta(days=1)
        
    return all_items


def run_pure_api_collector():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    master_sum_path = os.path.join(OUTPUT_DIR, "amberlyn_MASTER_Osszegzes.xlsx")
    master_raw_path = os.path.join(OUTPUT_DIR, "amberlyn_MASTER_Nyers.xlsx")

    print("[0. LÉPÉS] KORÁBBI MASTER NYERS FÁJL BEOLVASÁSA A DRIVE-RÓL...")
    download_master_file_from_drive("amberlyn_MASTER_Nyers.xlsx", GOOGLE_DRIVE_FOLDER_ID, master_raw_path)

    captured_token = None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        def handle_request(request):
            nonlocal captured_token
            auth_header = request.headers.get("authorization", "")
            if auth_header.startswith("Bearer ") and not captured_token:
                captured_token = auth_header

        page.on("request", handle_request)
        print("\n[1. LÉPÉS] Bejelentkezés a munkamenet-token megszerzéséhez...")
        page.goto(LOGIN_URL)
        page.wait_for_selector("#username", timeout=15000)
        page.locator("#username").fill(TEYA_EMAIL)
        page.press("#username", "Enter")

        password_selector = 'input[type="password"], input[name="password"]'
        page.wait_for_selector(password_selector, timeout=15000)
        page.locator(password_selector).fill(TEYA_PASSWORD)
        page.press(password_selector, "Enter")

        page.goto("https://business.teya.com/settlements")
        try:
            page.wait_for_selector('[data-testid="settlements-calendar"]', timeout=25000)
        except Exception:
            page.wait_for_timeout(3000)
        browser.close()

    if not captured_token:
        print("[SÚLYOS HIBA] Nem sikerült elcsípni a munkamenet-tokent.")
        return

    headers = {
        "accept": "*/*",
        "authorization": captured_token,
        "content-type": "application/json",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/151.0.0.0",
    }

    # IDŐSZAK: Utolsó 14 nap (2 hét) a napi kifizetésekhez
    end_date = datetime.now()
    start_date = end_date - timedelta(days=DAYS_BACK)

    print(f"\n[2. LÉPÉS] Adatok lekérése az elmúlt {DAYS_BACK} napra ({start_date.strftime('%Y-%m-%d')} -> {end_date.strftime('%Y-%m-%d')})...")
    data_items = fetch_all_settlements_in_chunks(start_date, end_date, headers)
    
    processed_payouts = set()

    print("\n[3. LÉPÉS] Tételek frissítése és rögzítése...")
    for item in data_items:
        if item.get("state") != "PAID": continue
        payout_id = item.get("payout_id")
        payout_date = item.get("settled_at") or item.get("payout_date", "")[:10]
        if not payout_date or payout_id in processed_payouts: continue

        processed_payouts.add(payout_id)
        print(f"\n-> Feldolgozás: {payout_date}")

        fees = item.get("fees", 0.0)
        summary_data = {
            "kifizetes": format_eur(item.get("amount")),
            "kezdeti_egyenleg": format_eur(item.get("initial_balance")),
            "tranzakciok": format_eur(item.get("sales")),
            "visszateritesek": format_eur(item.get("refunds")),
            "chargeback": format_eur(item.get("chargebacks")),
            "dijak": format_eur(-fees if fees > 0 else fees),
        }

        raw_file_name = f"amberlyn_{payout_date}.xlsx"
        summary_file_name = f"amberlyn_{payout_date}_Osszegzes.xlsx"
        raw_path = os.path.join(OUTPUT_DIR, raw_file_name)
        sum_path = os.path.join(OUTPUT_DIR, summary_file_name)

        create_summary_excel(sum_path, summary_data, payout_date)

        if payout_id:
            if generate_and_download_raw_excel(payout_id, raw_path, headers):
                append_to_master_raw(master_raw_path, raw_path, payout_date)

        upload_to_google_drive(sum_path, summary_file_name, GOOGLE_DRIVE_FOLDER_ID)
        upload_to_google_drive(raw_path, raw_file_name, GOOGLE_DRIVE_FOLDER_ID)

    # 4. A Master Összegző fájl teljes újraépítése a Drive-on lévő ÖSSZES napi összegző alapján
    rebuild_master_summary_from_all_files(master_sum_path, GOOGLE_DRIVE_FOLDER_ID)

    # Master fájlok szinkronizálása a Google Drive-ra
    print(f"\n[5. LÉPÉS] MASTER fájlok szinkronizálása a Google Drive-ra...")
    if os.path.exists(master_sum_path):
        upload_to_google_drive(master_sum_path, "amberlyn_MASTER_Osszegzes.xlsx", GOOGLE_DRIVE_FOLDER_ID)
    if os.path.exists(master_raw_path):
        upload_to_google_drive(master_raw_path, "amberlyn_MASTER_Nyers.xlsx", GOOGLE_DRIVE_FOLDER_ID)

    print("\n--> MINTA SIKERESEN LEFUTOTT!")


if __name__ == "__main__":
    run_pure_api_collector()
